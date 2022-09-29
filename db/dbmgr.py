# +------------+--------------+-----------------------------------------------------------+
# |   Author   |     Date     |                         Changed                           |
# +------------+--------------+-----------------------------------------------------------+
# |  Andrew A. |  2022/09/29  | Merged dbmgr_mainwindow.py and dbmgr_pathselect.py        |
# +------------+--------------+-----------------------------------------------------------+
# |  Andrew A. |  2022/09/29  | Implemented basic features                                |
# +------------+--------------+-----------------------------------------------------------+
# |  Andrew A. |  2022/09/29  | Added export feature                                      |
# +------------+--------------+-----------------------------------------------------------+

from PyQt5.QtCore import QThread, pyqtSignal, pyqtSlot, Qt
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import *
import matplotlib.pyplot as plt
import openpyxl
import sqlite3
import sys


class PathSelectWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi()

    def setupUi(self):
        font = QtGui.QFont()
        font.setFamily("Apple SD Gothic Neo")

        self.setWindowTitle("온도미쁨 - 데이터베이스 관리 도구")
        self.resize(821, 57)

        self.centralwidget = QWidget(self)
        self.horizontalLayout_2 = QHBoxLayout(self.centralwidget)

        self.baseHori = QHBoxLayout()

        self.pathEdit = QLineEdit("/Users/flyahn06/code/OndoMippum/resources/main.db", self.centralwidget)
        self.pathEdit.setFont(font)
        self.baseHori.addWidget(self.pathEdit)

        self.loadBtn = QPushButton("불러오기", self.centralwidget)
        self.loadBtn.setFont(font)
        self.loadBtn.clicked.connect(self.loadDB)

        self.baseHori.addWidget(self.loadBtn)
        self.horizontalLayout_2.addLayout(self.baseHori)
        self.setCentralWidget(self.centralwidget)

        self.show()
        # TODO: For debugging
        self.loadDB()

    def loadDB(self):
        path = self.pathEdit.text()

        try:
            open(path, 'rb')
        except FileNotFoundError:
            return

        self.next = MainWindow(path)
        self.close()


class filePathDialog(QDialog):
    path = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setupUi()

    def setupUi(self):
        self.resize(635, 89)
        self.setMinimumSize(QtCore.QSize(635, 89))
        self.setMaximumSize(QtCore.QSize(10000, 89))
        self.setWindowTitle("온도미쁨 - 데이터베이스 관리 도구")

        self.verticalLayout_2 = QVBoxLayout(self)
        self.base = QVBoxLayout()
        self.base_in = QHBoxLayout()

        self.filePathLbl = QLabel(self)
        font = QtGui.QFont()
        font.setFamily("Apple SD Gothic Neo")
        self.filePathLbl.setFont(font)
        self.base_in.addWidget(self.filePathLbl)

        self.filePathEdit = QLineEdit(self)
        self.base_in.addWidget(self.filePathEdit)
        self.base.addLayout(self.base_in)

        self.verticalLayout_2.addLayout(self.base)
        self.buttonBox = QDialogButtonBox(self)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QDialogButtonBox.Cancel|QDialogButtonBox.Ok)
        self.verticalLayout_2.addWidget(self.buttonBox)

        self.buttonBox.accepted.connect(self.accept)  # type: ignore
        self.buttonBox.rejected.connect(self.reject)  # type: ignore

        self.show()

    def accept(self):
        print("accepted", self.filePathEdit.text())
        self.path.emit(self.filePathEdit.text())
        self.destroy()


    def reject(self):
        print("rejected")
        self.destroy()



class DBLoadWorker(QThread):
    data = pyqtSignal(list)
    error = pyqtSignal(Exception)

    def __init__(self, cur):
        super().__init__()
        self.cur = cur

    def run(self):
        while True:
            self.msleep(100)

    def fetch(self, name="", code=0, time_from="", time_to=""):
        query = "SELECT * FROM LOG "

        if name:
            query += f'WHERE name="{name}"'

        if code:
            query += "WHERE classcode=" + str(code)

        # TODO: 시간 검색 구현
        print(query)
        return self.cur.execute(query).fetchall()

class Grapher(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        while True:
            self.msleep(100)

    def graph(self, data, name):
        plt.title(name)
        plt.plot(data[0], data[1], label='체온')
        plt.legend()
        plt.show()

class ExportWorker(QThread):
    def __init__(self):
        super().__init__()

    def run(self):
        while True:
            self.msleep(100)

    def export(self, data, path):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "10912 안동기"
        ws['A1'] = "번호"
        ws['B1'] = "날짜"
        ws['C1'] = "학번"
        ws['D1'] = "이름"
        ws['E1'] = "체온"

        depth = len(data) + 1
        start = "A2"
        end = "E" + str(depth)
        cell_range = start + ":" + end

        print(cell_range)

        for row, datum in zip(ws[start:end], data):
            for cell, line in zip(row, datum):
                cell.value = line

        wb.save(path)

class MainWindow(QMainWindow):
    def __init__(self, path):
        super().__init__()
        self.con = sqlite3.connect(path)
        self.cur = self.con.cursor()
        self.res = []
        self.setupUi()

    def setupUi(self):
        font = QtGui.QFont()
        font.setFamily("Apple SD Gothic Neo")

        self.resize(898, 578)
        self.centralwidget = QWidget(self)
        self.setWindowTitle("온도미쁨 - 데이터베이스 관리 도구")

        self.verticalLayout_2 = QVBoxLayout(self.centralwidget)
        self.baseVert = QVBoxLayout()
        self.upperTimeSearchHori = QHBoxLayout()

        self.timeSearchEnableChkBox = QCheckBox("시간 검색", self.centralwidget)
        self.timeSearchEnableChkBox.setFont(font)
        self.timeSearchEnableChkBox.clicked.connect(self.enableTimeSearch)
        self.upperTimeSearchHori.addWidget(self.timeSearchEnableChkBox)

        self.timeSearchFromEdit = QDateTimeEdit(self.centralwidget)
        self.timeSearchFromEdit.setFont(font)
        self.timeSearchFromEdit.setEnabled(False)
        self.upperTimeSearchHori.addWidget(self.timeSearchFromEdit)

        self.timeSearchFromLbl = QLabel("부터", self.centralwidget)
        self.timeSearchFromLbl.setFont(font)
        self.timeSearchFromLbl.setEnabled(False)
        self.upperTimeSearchHori.addWidget(self.timeSearchFromLbl)

        self.timeSearchToEdit = QDateTimeEdit(self.centralwidget)
        self.timeSearchToEdit.setFont(font)
        self.timeSearchToEdit.setEnabled(False)
        self.upperTimeSearchHori.addWidget(self.timeSearchToEdit)

        self.timeSearchToLbl = QLabel("까지", self.centralwidget)
        self.timeSearchToLbl.setFont(font)
        self.timeSearchToLbl.setEnabled((False))
        self.upperTimeSearchHori.addWidget(self.timeSearchToLbl)
        self.baseVert.addLayout(self.upperTimeSearchHori)

        self.upperNameSearchHori = QHBoxLayout()

        self.nameSearchEdit = QLineEdit(self.centralwidget)
        self.nameSearchEdit.setFont(font)
        self.nameSearchEdit.setPlaceholderText("검색할 대상의 이름을 입력하세요. 입력하지 않으면 모두(*) 가져옵니다.")
        self.upperNameSearchHori.addWidget(self.nameSearchEdit)

        self.searchBtn = QPushButton("검색", self.centralwidget)
        self.searchBtn.setFont(font)
        self.searchBtn.clicked.connect(self.getData)
        self.upperNameSearchHori.addWidget(self.searchBtn)

        self.baseVert.addLayout(self.upperNameSearchHori)

        self.tableWidget = QTableWidget(self.centralwidget)
        self.baseVert.addWidget(self.tableWidget)

        self.lowerExportGraphHori = QHBoxLayout()

        self.graphBtn = QPushButton("그래프 보기", self.centralwidget)
        self.graphBtn.setFont(font)
        self.graphBtn.clicked.connect(self.graph)
        self.lowerExportGraphHori.addWidget(self.graphBtn)

        self.exportBtn = QPushButton("액셀 파일로 내보내기", self.centralwidget)
        self.exportBtn.setFont(font)
        self.exportBtn.clicked.connect(self.export)
        self.lowerExportGraphHori.addWidget(self.exportBtn)

        self.baseVert.addLayout(self.lowerExportGraphHori)
        self.verticalLayout_2.addLayout(self.baseVert)
        self.setCentralWidget(self.centralwidget)

        self.statusbar = QStatusBar(self)
        self.setStatusBar(self.statusbar)

        self.db = DBLoadWorker(self.cur)
        self.db.start()

        self.grapher = Grapher()
        self.grapher.start()

        self.exporter = ExportWorker()
        self.exporter.start()

        self.show()

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Return or e.key() == Qt.Key_Enter:
            self.getData()

    def enableTimeSearch(self):
        if self.timeSearchEnableChkBox.isChecked():
            self.timeSearchFromLbl.setEnabled(True)
            self.timeSearchToLbl.setEnabled(True)
            self.timeSearchFromEdit.setEnabled(True)
            self.timeSearchToEdit.setEnabled(True)
        else:
            self.timeSearchFromLbl.setEnabled(False)
            self.timeSearchToLbl.setEnabled(False)
            self.timeSearchFromEdit.setEnabled(False)
            self.timeSearchToEdit.setEnabled(False)

    def getData(self):
        if self.nameSearchEdit.text():
            condition = self.nameSearchEdit.text()

            if condition.isdigit():
                res = self.db.fetch(code=int(condition))
            else:
                res = self.db.fetch(name=condition)
        else:
            res = self.db.fetch()

        self.res = res
        self.tableWidget.setRowCount(len(res))
        self.tableWidget.setColumnCount(4)

        for index, packed_data in enumerate(res):
            ID, date, classcode, name, temp = packed_data
            self.tableWidget.setItem(index, 0, QTableWidgetItem(date))
            self.tableWidget.setItem(index, 1, QTableWidgetItem(str(classcode)))
            self.tableWidget.setItem(index, 2, QTableWidgetItem(name))
            self.tableWidget.setItem(index, 3, QTableWidgetItem(str(temp)))

        column_headers = ("시간", "학번", "이름", "체온")
        self.tableWidget.setHorizontalHeaderLabels(column_headers)

    def graph(self):
        if not self.nameSearchEdit.text():
            QMessageBox.critical(self, "오류", "체온에 대한 그래프는 한 사람만 선택되었을 때 그릴 수 있습니다. \n이름/학번 검색 후 다시 시도해주세요",
                                 QMessageBox.Yes, QMessageBox.Yes)
            return

        if not self.res:
            QMessageBox.critical(self, "오류", "검색된 결과가 0건입니다.",
                                 QMessageBox.Yes, QMessageBox.Yes)
            return

        self.grapher.graph(([x[1] for x in self.res], [x[4] for x in self.res]), str(self.res[0][2]) + self.res[0][3])

    def export(self):
        if not self.res:
            return

        a = filePathDialog()
        a.path.connect(self.export_helper)
        self.__next__ = a

    @pyqtSlot(str)
    def export_helper(self, path):
        print("signal", path)
        self.exporter.export(self.res, path)


app = QApplication(sys.argv)
pathselect = PathSelectWindow()
sys.exit(app.exec_())


